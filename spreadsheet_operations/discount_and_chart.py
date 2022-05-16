import openpyxl as xl
from openpyxl.chart import BarChart, Reference


class ProcessWorkbook:

    def __init__(self, filename, sheet_number):
        self.filename = filename
        self.sheet_number = sheet_number
        self.wb = xl.load_workbook(self.filename)
        self.sheet = self.wb[self.sheet_number]

    def add_discounted_column(self, discount, column, start_row):
        for row in range(start_row, self.sheet.max_row + 1):
            cell = self.sheet.cell(row, column)
            corrected_price = cell.value * discount
            corrected_price_cell = self.sheet.cell(row, column + 1)
            corrected_price_cell.value = corrected_price
            self.wb.save(self.filename)

    def add_bar_chart(self, first_row_of_data, data_column, place_of_chart):

        values = Reference(self.sheet,
                           min_row=first_row_of_data, max_row=self.sheet.max_row,
                           min_col=data_column,
                           max_col=data_column)

        chart = BarChart()
        chart.add_data(values)
        self.sheet.add_chart(chart, place_of_chart)

        self.wb.save(self. filename)
